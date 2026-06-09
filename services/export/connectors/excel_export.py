"""Excel書式付きエクスポーター: GraphModelをExcelファイルに出力

仕様書 windows-integration.md Phase W-1/W-2 に基づく実装:
- W-1: テーブルデータ書き出し（メイリオ書式・ヘッダー色・列幅自動調整）
- W-2: 配列データ書き出し（複数シート・Summaryシート付き）

openpyxlを使用し、クロスプラットフォームで動作する。

[READMEへ戻る](../../../README.md)
"""

from __future__ import annotations

import io
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from jj_types import GraphModel
from plugins.base.exporter import AbstractExporter

# ================================================================
# スタイル定数
# ================================================================

FONT_NAME = "メイリオ"
FONT_SIZE = 10
HEADER_BG_COLOR = "4472C4"

BODY_FONT = Font(name=FONT_NAME, size=FONT_SIZE)
HEADER_FONT = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="top")


# ================================================================
# エクスポーター
# ================================================================


class ExcelExporter(AbstractExporter):
    """Excel形式エクスポーター（書式付き）

    テーブルデータを書式付きExcelファイルとして出力する。
    """

    format = "xlsx"
    priority = 12

    def export(self, graph: GraphModel, **kwargs: Any) -> dict[str, Any]:
        """Excelエクスポート

        kwargs:
            project_root: Path - プロジェクトルート
            nodes: list[Node] | None - 事前選択済みノードリスト
            type_filter: str | None - ノードタイプフィルタ
            output_file: str | None - 出力ファイル名
            sheet_name: str - シート名（デフォルト: "Data"）
            flatten: bool - プロパティ平坦化（デフォルト: True）
        """
        from services.export.connectors.csv_json import flatten_properties

        project_root = kwargs.get("project_root", Path("."))
        nodes = kwargs.get("nodes")
        type_filter = kwargs.get("type_filter")
        output_file = kwargs.get("output_file")
        sheet_name = kwargs.get("sheet_name", "Data")
        flatten = kwargs.get("flatten", True)

        # ノードのフィルタリング
        if nodes is not None:
            filtered_nodes = list(nodes)
        else:
            filtered_nodes = list(graph.nodes)
            if type_filter:
                filtered_nodes = [n for n in filtered_nodes if n.type == type_filter]

        if not filtered_nodes:
            raise ValueError("対象ノードが見つかりません。")

        # データ行の構築
        data_rows: list[dict[str, Any]] = []
        for node in filtered_nodes:
            base: dict[str, Any] = {
                "name": node.name,
                "type": node.type,
                "format": node.format,
            }
            if flatten:
                props = flatten_properties(node.properties)
            else:
                props = dict(node.properties)
            base.update(props)
            data_rows.append(base)

        # カラム収集
        all_keys: list[str] = ["name", "type", "format"]
        seen_keys: set[str] = set(all_keys)
        for row in data_rows:
            for key in row:
                if key not in seen_keys:
                    all_keys.append(key)
                    seen_keys.add(key)

        # ワークブック生成
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        _write_table_sheet(ws, all_keys, data_rows)

        # 出力
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"table_{timestamp}.xlsx"
        output_path = project_root / ".j2" / "exports" / output_file
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb.save(str(output_path))
        return {"output_path": output_path, "count": len(data_rows)}

    def format_cli_result(self, result: dict[str, Any], project_root: Path) -> str:
        output_path = result.get("output_path")
        count = result.get("count", 0)
        if output_path:
            try:
                rel = Path(output_path).relative_to(project_root)
            except ValueError:
                rel = output_path
            return f"Excelエクスポート完了: {rel} ({count}件)"
        return f"Excelエクスポート完了 ({count}件)"


# ================================================================
# テーブルデータ書き出し（W-1）
# ================================================================


def export_table_to_excel(
    df: Any,
    output_dir: Path,
    filename_prefix: str = "table",
    sheet_name: str = "Data",
    freeze_panes: tuple[int, int] = (1, 0),
) -> Path:
    """DataFrameをExcelファイルに書式付きで書き出す

    Features:
    - フォント: メイリオ 10pt
    - ヘッダー行: 太字、背景色 #4472C4、文字色 白
    - 列幅: 内容に基づく自動調整（日本語は2文字幅換算）
    - ウィンドウ枠固定: ヘッダー行

    Args:
        df: pandas DataFrame
        output_dir: 出力ディレクトリ
        filename_prefix: ファイル名の接頭辞
        sheet_name: シート名
        freeze_panes: ウィンドウ枠固定位置 (row, col) - 1-indexed

    Returns:
        出力ファイルパス
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    columns = list(df.columns)
    rows = df.values.tolist()

    # データ辞書のリストに変換
    data_rows = [{col: row[i] for i, col in enumerate(columns)} for row in rows]
    _write_table_sheet(ws, columns, data_rows)

    # ウィンドウ枠固定
    if freeze_panes[0] > 0 or freeze_panes[1] > 0:
        freeze_row = freeze_panes[0] + 1  # openpyxl は 1-indexed
        freeze_col = freeze_panes[1] + 1
        ws.freeze_panes = ws.cell(row=freeze_row, column=freeze_col)

    # 出力
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"{filename_prefix}_{timestamp}.xlsx"
    output_dir.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return output_path


def export_table_to_excel_bytes(
    df: Any,
    sheet_name: str = "Data",
) -> bytes:
    """DataFrameを書式付きExcelのバイト列に変換する

    Streamlitのダウンロードボタン用。

    Args:
        df: pandas DataFrame
        sheet_name: シート名

    Returns:
        Excelファイルのバイト列
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    columns = list(df.columns)
    rows = df.values.tolist()

    data_rows = [{col: row[i] for i, col in enumerate(columns)} for row in rows]
    _write_table_sheet(ws, columns, data_rows)

    # ウィンドウ枠固定（ヘッダー行）
    ws.freeze_panes = ws.cell(row=2, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ================================================================
# 配列データ書き出し（W-2）
# ================================================================


def export_array_data_to_excel(
    array_data: list[dict[str, Any]],
    output_dir: Path,
    filename_prefix: str = "array_data",
) -> Path:
    """配列データをExcelに書き出す

    シート構成:
    - "Summary" シート: ノード名、プロパティ一覧
    - 各ノードごとにシート: x, y 列 + メタデータヘッダー

    Args:
        array_data: [{"name": str, "x": list, "y": list, "props": dict}, ...]
        output_dir: 出力ディレクトリ
        filename_prefix: ファイル名の接頭辞

    Returns:
        出力ファイルパス
    """
    wb = Workbook()

    # Summary シート
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_headers = ["No.", "Name"]
    # プロパティキーを収集
    all_prop_keys: list[str] = []
    seen_prop_keys: set[str] = set()
    for item in array_data:
        for key in item.get("props", {}):
            if key not in seen_prop_keys:
                all_prop_keys.append(key)
                seen_prop_keys.add(key)

    summary_headers.extend(all_prop_keys)
    summary_headers.extend(["Data Points", "X Range", "Y Range"])

    summary_rows: list[dict[str, Any]] = []
    for i, item in enumerate(array_data, 1):
        row: dict[str, Any] = {"No.": i, "Name": item.get("name", "")}
        for key in all_prop_keys:
            row[key] = item.get("props", {}).get(key, "")

        x_data = item.get("x", [])
        y_data = item.get("y", [])
        row["Data Points"] = len(x_data)
        if x_data:
            row["X Range"] = f"{min(x_data):.4g} ~ {max(x_data):.4g}"
        if y_data:
            row["Y Range"] = f"{min(y_data):.4g} ~ {max(y_data):.4g}"

        summary_rows.append(row)

    _write_table_sheet(ws_summary, summary_headers, summary_rows)

    # 各データシート
    for i, item in enumerate(array_data, 1):
        name = item.get("name", f"Data_{i}")
        # シート名は31文字制限 + 特殊文字除去
        safe_name = _sanitize_sheet_name(name, i)
        ws = wb.create_sheet(title=safe_name)

        # メタデータヘッダー（行1-3）
        props = item.get("props", {})
        ws.cell(row=1, column=1, value="Name").font = BODY_FONT
        ws.cell(row=1, column=2, value=name).font = BODY_FONT

        row_offset = 2
        for key, value in props.items():
            ws.cell(row=row_offset, column=1, value=key).font = BODY_FONT
            ws.cell(row=row_offset, column=2, value=str(value)).font = BODY_FONT
            row_offset += 1

        # 空行
        row_offset += 1

        # データヘッダー
        x_label = item.get("x_label", "X")
        y_label = item.get("y_label", "Y")
        for col_idx, header in enumerate([x_label, y_label], 1):
            cell = ws.cell(row=row_offset, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT

        # データ行
        x_data = item.get("x", [])
        y_data = item.get("y", [])
        for j, (x_val, y_val) in enumerate(zip(x_data, y_data, strict=False)):
            ws.cell(row=row_offset + 1 + j, column=1, value=x_val).font = BODY_FONT
            ws.cell(row=row_offset + 1 + j, column=2, value=y_val).font = BODY_FONT

        # 列幅調整
        _auto_column_width(ws)

    # 出力
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"{filename_prefix}_{timestamp}.xlsx"
    output_dir.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return output_path


def export_array_data_to_excel_bytes(
    array_data: list[dict[str, Any]],
) -> bytes:
    """配列データをExcelバイト列に変換する（Streamlitダウンロード用）

    Args:
        array_data: [{"name": str, "x": list, "y": list, "props": dict}, ...]

    Returns:
        Excelファイルのバイト列
    """
    import tempfile

    with tempfile.TemporaryDirectory() as tmpdir:
        path = export_array_data_to_excel(array_data, Path(tmpdir), "download")
        return path.read_bytes()


# ================================================================
# ヘルパー関数
# ================================================================


def _write_table_sheet(
    ws: Any,
    columns: list[str],
    data_rows: list[dict[str, Any]],
) -> None:
    """シートにテーブルデータを書式付きで書き込む

    Args:
        ws: openpyxl Worksheet
        columns: カラム名リスト
        data_rows: データ辞書のリスト
    """
    # ヘッダー行
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    # データ行
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name)
            # リスト/辞書はJSON文字列化
            if isinstance(value, (list, dict)):
                import json

                value = json.dumps(value, ensure_ascii=False)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGNMENT

    # 列幅自動調整
    _auto_column_width(ws)

    # ウィンドウ枠固定（ヘッダー行）
    ws.freeze_panes = ws.cell(row=2, column=1)


def _auto_column_width(ws: Any) -> None:
    """列幅を内容に基づいて自動調整する

    日本語文字は2文字幅として計算する。
    """
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)

        for cell in col_cells:
            if cell.value is not None:
                text = str(cell.value)
                # 日本語文字は2文字幅
                length = sum(2 if unicodedata.east_asian_width(c) in ("F", "W") else 1 for c in text)
                max_length = max(max_length, length)

        # 最小幅 8、最大幅 50
        adjusted_width = min(max(max_length + 2, 8), 50)
        ws.column_dimensions[col_letter].width = adjusted_width


def _sanitize_sheet_name(name: str, index: int) -> str:
    """シート名をExcelの制約に適合させる

    - 31文字以内
    - 特殊文字 []:*?/\\ を除去
    """
    # 特殊文字除去
    for char in r"[]:*?/\\":
        name = name.replace(char, "")
    # 31文字制限
    if len(name) > 28:
        name = f"{name[:25]}_{index}"
    return name or f"Sheet_{index}"
