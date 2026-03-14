"""XLSXメタデータパーサー: .xlsxファイルからメタデータを抽出

openpyxlライブラリを使用して以下の情報を抽出する:
- シート名一覧
- 各シートの行数・列数
- コアプロパティ（作成者・タイトル等）

[READMEへ戻る](../../../../README.md)
"""

from __future__ import annotations

import logging
import sys
from typing import TYPE_CHECKING, Any

from services.parse.base import AbstractFileParser

if TYPE_CHECKING:
    from services.graph.project_graph import ProjectGraph

logger = logging.getLogger(__name__)


class XlsxMetadataParser(AbstractFileParser):
    """XLSXファイルのメタデータを抽出してノードプロパティに付与する

    priority=35: 入出力関係解析(30-33)の後、ディレクトリ関係(50)の前に実行
    """

    priority = 35

    def apply(self, graph: ProjectGraph) -> ProjectGraph:
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.debug("openpyxl が未インストールのため XLSX パーサーをスキップ")
            return graph

        for node in graph.nodes:
            if node.format != "xlsx":
                continue

            file_path = graph.project_root / node.properties.get("path", "")
            if not file_path.exists():
                continue

            try:
                metadata = _extract_xlsx_metadata(file_path, load_workbook)
                node.properties.update(metadata)
            except Exception as e:
                print(
                    f"警告: XLSX メタデータ抽出に失敗 ({file_path.name}): {e}",
                    file=sys.stderr,
                )

        return graph


def _extract_xlsx_metadata(file_path: Any, load_workbook: Any) -> dict[str, Any]:
    """XLSXファイルからメタデータを抽出する

    Args:
        file_path: XLSXファイルのパス
        load_workbook: openpyxl の load_workbook 関数

    Returns:
        抽出されたメタデータの辞書
    """
    # read_only=True + data_only=True で高速読み込み
    wb = load_workbook(str(file_path), read_only=True, data_only=True)

    metadata: dict[str, Any] = {}

    try:
        # シート名一覧
        sheet_names = wb.sheetnames
        metadata["xlsx.sheet_count"] = len(sheet_names)
        metadata["xlsx.sheet_names"] = sheet_names

        # 各シートの行数・列数
        sheet_dimensions: dict[str, dict[str, int]] = {}
        total_rows = 0

        for name in sheet_names:
            ws = wb[name]
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            sheet_dimensions[name] = {"rows": max_row, "columns": max_col}
            total_rows += max_row

        metadata["xlsx.total_rows"] = total_rows
        metadata["xlsx.sheet_dimensions"] = sheet_dimensions

        # コアプロパティ
        props = wb.properties
        if props:
            if props.creator:
                metadata["xlsx.author"] = props.creator
            if props.title:
                metadata["xlsx.title"] = props.title
            if props.subject:
                metadata["xlsx.subject"] = props.subject
            if props.created:
                metadata["xlsx.created"] = str(props.created)
            if props.modified:
                metadata["xlsx.modified"] = str(props.modified)
    finally:
        wb.close()

    return metadata
